
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection # 글꼴 테두리 맞춤 채우기 보호

from openpyxl import load_workbook

def FixingFile(path, dc_rate):
    wb = load_workbook(path)
    ws = wb['Sheet1']
    ws.column_dimensions['A'].width = 11.0
    ws.column_dimensions['B'].width = 12.0
    ws.column_dimensions['C'].width = 8.0
    
    temp_sum = 0
    col_C = ws['C']
    for cell in col_C:
        if cell.value == 'price':
            continue
        if cell.value == None:
            continue
        cell.number_format = '#,###'
        temp_sum += int(cell.value) 

    ws.cell(row = 2, column = 5, value = '합계금액')
    ws.cell(row = 2, column = 6, value = temp_sum).number_format = '#,###'

    if dc_rate is not None and int(dc_rate) > 0:
        dc_rate = int(dc_rate) / 100
        ws.cell(row = 3, column = 5, value = '할인금액({}%)'.format(100 * dc_rate))
        ws.cell(row = 3, column = 6, value = temp_sum * (1 - dc_rate)).number_format = '#,###'
    row1 = ws[1]
    for cell in row1:
        cell.font = Font(bold = True)

    ws.cell(row = 2, column = 12).font = Font(bold = True)
    ws.cell(row = 2, column = 13).font = Font(bold = True)
    ws.cell(row = 3, column = 12).font = Font(bold = True, color = '00FF0000')
    ws.cell(row = 3, column = 13).font = Font(bold = True, color = '00FF0000')
    #row1.font = Font(bold = True)

    wb.save(path)

